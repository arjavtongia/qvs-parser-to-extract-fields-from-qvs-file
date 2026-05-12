"""
QVD to QVW Mapper
-----------------
Reads QVD paths from column A of the "QVD List" sheet in an Excel workbook,
extracts keywords (SNOW, TAI, TMD, Stability, etc.), searches for matching .QVW
files in a given root directory, then scans each QVW for .qvd references.
Matching QVW paths are written to column C.

Usage:
    python qvd_qvw_mapper.py

Configure the three variables in the CONFIG section below before running.
"""

import os
import re
import glob
from openpyxl import load_workbook

# ─── CONFIG ──────────────────────────────────────────────────────────────────

EXCEL_PATH = r"C:\path\to\your\Stability_Level_2.xlsx"   # Full path to your workbook
QVW_ROOT   = r"\\stor223ncs1.london.ms.com\s270960\arjavt\My Documents\Model"  # Root folder to search QVWs
SHEET_NAME = "QVD List"

# Keywords to detect and the glob pattern to find matching QVW files.
# Add more entries here if needed.
KEYWORD_PATTERNS = {
    "SNOW":      "*SNOW*.QVW",
    "TAI":       "*TAI*.QVW",
    "TMD":       "*TMD*.QVW",
    "STABILITY": "*STABILITY*.QVW",
    "ZITI":      "*ZITI*.QVW",
    "QAPM":      "*QAPM*.QVW",
}

# ─────────────────────────────────────────────────────────────────────────────


def extract_keyword(qvd_path: str) -> str | None:
    """
    Pull the keyword out of a QVD path like:
      $(vQVDExtractPath)SNOWvEX_11797_QAPM_ClClass.qvd
    Returns the uppercased keyword or None if nothing matched.
    """
    upper = qvd_path.upper()
    for kw in KEYWORD_PATTERNS:
        if kw in upper:
            return kw
    return None


def find_qvw_files(keyword: str) -> list[str]:
    """Return all QVW file paths whose name matches the keyword pattern."""
    pattern = KEYWORD_PATTERNS[keyword]
    # case-insensitive glob on Windows; on Linux you may need a custom walk
    matches = glob.glob(os.path.join(QVW_ROOT, "**", pattern), recursive=True)
    # also try the root directly
    matches += glob.glob(os.path.join(QVW_ROOT, pattern))
    # exclude backup/copy files
    matches = [
        m for m in matches
        if not any(x in os.path.basename(m).upper() for x in ("BCK", "COPY"))
    ]
    return list(set(matches))


def qvd_refs_in_file(qvw_path: str) -> list[str]:
    """
    Open a QVW (binary but contains readable text sections) and extract
    every .qvd reference found between the keyword and '.qvd'.
    Returns the raw matched strings (the substring between keyword\ and .qvd).
    """
    refs = []
    try:
        # QVW files are binary; read as latin-1 to avoid decode errors
        with open(qvw_path, "r", encoding="latin-1", errors="replace") as f:
            content = f.read()
        # Match anything ending in .qvd (case-insensitive)
        hits = re.findall(r'[A-Za-z0-9_\$\(\)\\\/\-\.]+\.qvd', content, re.IGNORECASE)
        refs = list(set(hits))
    except Exception as e:
        print(f"  [WARN] Could not read {qvw_path}: {e}")
    return refs


def qvw_references_qvd(qvw_path: str, qvd_path: str) -> bool:
    """
    Check if the QVW file contains a reference to the specific QVD.
    We match on the filename part of the QVD path (last segment).
    """
    qvd_filename = os.path.basename(qvd_path).lower()
    # Strip the variable prefix like $(vQVDExtractPath) to get just the filename
    # e.g. "$(vQVDExtractPath)SNOWvEX_11797_QAPM_ClClass.qvd" -> "snowvex_11797_qapm_clclass.qvd"
    try:
        with open(qvw_path, "r", encoding="latin-1", errors="replace") as f:
            content = f.read().lower()
        return qvd_filename in content
    except Exception as e:
        print(f"  [WARN] Could not read {qvw_path}: {e}")
        return False


def main():
    print(f"Loading workbook: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    print(f"Processing sheet: {SHEET_NAME}")

    # Cache QVW file lists per keyword to avoid repeated disk scans
    qvw_cache: dict[str, list[str]] = {}

    updated = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=False), start=2):
        cell_a = row[0]
        qvd_path = cell_a.value
        if not qvd_path or not str(qvd_path).strip():
            continue

        qvd_path = str(qvd_path).strip()
        keyword = extract_keyword(qvd_path)
        if not keyword:
            print(f"  Row {row_idx}: No keyword match for: {qvd_path}")
            continue

        # Find QVW files for this keyword
        if keyword not in qvw_cache:
            qvw_cache[keyword] = find_qvw_files(keyword)
            print(f"  Found {len(qvw_cache[keyword])} QVW(s) for keyword '{keyword}'")

        qvw_files = qvw_cache[keyword]
        matching_qvws = []

        for qvw_path in qvw_files:
            if qvw_references_qvd(qvw_path, qvd_path):
                matching_qvws.append(qvw_path)

        if matching_qvws:
            # Write all matching paths to column C, semicolon-separated if multiple
            result = "; ".join(matching_qvws)
            ws.cell(row=row_idx, column=3).value = result
            print(f"  Row {row_idx}: {os.path.basename(qvd_path)} -> {len(matching_qvws)} match(es)")
            updated += 1
        else:
            print(f"  Row {row_idx}: {os.path.basename(qvd_path)} -> no QVW match found")

    wb.save(EXCEL_PATH)
    print(f"\nDone. Updated {updated} rows. Saved to: {EXCEL_PATH}")


if __name__ == "__main__":
    main()
