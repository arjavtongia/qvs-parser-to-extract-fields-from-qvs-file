"""
qvs_extractor.py
----------------
Extracts Tab Name, Table Name, Field, and FROM from a Qlik QVS script.

Rules:
  - Table Name : the label before the colon (e.g. "OrgStructure"). Never a field expression.
  - Field      : one row per field (full expression preserved).
  - FROM       : source name only. Qualifiers (ooxml/qvd/WHERE/GROUP BY) are stripped.
  - Commented lines (// ...) are fully ignored.

Usage:
    python qvs_extractor.py <path_to_qvs_file> [output_xlsx]
"""

import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ─── text helpers ─────────────────────────────────────────────────────────────

def strip_inline_comment(line):
    result, in_str, qchar, i = [], False, None, 0
    while i < len(line):
        ch = line[i]
        if in_str:
            result.append(ch)
            if ch == qchar:
                in_str = False
        else:
            if ch in ("'", '"'):
                in_str, qchar = True, ch
                result.append(ch)
            elif ch == '/' and i + 1 < len(line) and line[i+1] == '/':
                break
            else:
                result.append(ch)
        i += 1
    return ''.join(result)


def is_full_comment(line):
    return line.strip().startswith('//')


# ─── tokeniser ────────────────────────────────────────────────────────────────

_TAB_MARKER = re.compile(r'^\s*/{2,}\s*\$tab\s+(.+)', re.IGNORECASE)


def tokenise(script):
    current_tab = "Main"
    buf = []
    for raw in script.splitlines():
        m = _TAB_MARKER.match(raw)
        if m:
            current_tab = m.group(1).strip()
            continue
        if is_full_comment(raw):
            continue
        cleaned = strip_inline_comment(raw)
        if not cleaned.strip():
            continue
        buf.append(cleaned)
        joined = ' '.join(buf)
        while ';' in joined:
            idx = joined.index(';')
            stmt = joined[:idx].strip()
            if stmt:
                yield current_tab, stmt
            joined = joined[idx+1:].strip()
            buf = [joined] if joined else []
    if buf:
        stmt = ' '.join(buf).strip()
        if stmt:
            yield current_tab, stmt


# ─── field splitter (respects parens and strings) ────────────────────────────

def split_fields(raw):
    fields, depth, buf, in_str, qchar = [], 0, [], False, None
    for ch in raw:
        if in_str:
            buf.append(ch)
            if ch == qchar:
                in_str = False
        elif ch in ("'", '"'):
            in_str, qchar = True, ch
            buf.append(ch)
        elif ch == '(':
            depth += 1; buf.append(ch)
        elif ch == ')':
            depth -= 1; buf.append(ch)
        elif ch == ',' and depth == 0:
            val = ''.join(buf).strip().rstrip(',').strip()
            if val:
                fields.append(val)
            buf = []
        else:
            buf.append(ch)
    val = ''.join(buf).strip().rstrip(',').strip()
    if val:
        fields.append(val)
    return fields


# ─── FROM cleaner ─────────────────────────────────────────────────────────────

_FROM_STOP = re.compile(
    r'(?:'
    r'\s+WHERE\b'
    r'|\s+GROUP\s+BY\b'
    r'|\s+ORDER\s+BY\b'
    r'|\((?:ooxml|qvd|txt|csv|fix|dif|biff|html|xml|kml|svg)\b'
    r')',
    re.IGNORECASE,
)

def clean_from(raw):
    raw = re.sub(r'\s+', ' ', raw).strip()
    m = _FROM_STOP.search(raw)
    if m:
        raw = raw[:m.start()].strip()
    return raw.rstrip(';').strip()


# ─── prefix helpers ───────────────────────────────────────────────────────────

_ALL_PREFIXES = re.compile(
    r'^(?:(?:NOCONCATENATE|CONCATENATE|'
    r'INNER\s+JOIN|LEFT\s+JOIN|RIGHT\s+JOIN|JOIN|'
    r'INNER\s+KEEP|LEFT\s+KEEP|RIGHT\s+KEEP|KEEP)'
    r'\s*(?:\([\w\s\[\]]+\))?\s*)+',
    re.IGNORECASE,
)

def strip_prefixes(text):
    prev = None
    while prev != text:
        prev = text
        text = _ALL_PREFIXES.sub('', text).strip()
    return text


def extract_join_label(text):
    """Return e.g. 'LEFT JOIN(Temp)' from a prefix string, or ''."""
    m = re.match(
        r'^((?:INNER\s+JOIN|LEFT\s+JOIN|RIGHT\s+JOIN|JOIN|'
        r'INNER\s+KEEP|LEFT\s+KEEP|RIGHT\s+KEEP|KEEP)'
        r'\s*(?:\([\w\s\[\]]+\))?)',
        text.strip(), re.IGNORECASE,
    )
    return re.sub(r'\s+', ' ', m.group(1)).strip() if m else ''


# ─── core parser ──────────────────────────────────────────────────────────────
#
# Approach: split the normalised statement into up to 3 parts manually,
# so we never rely on greedy/lazy regex across the entire statement.
#
# Step 1 – detect and remove leading prefixes (NOCONCATENATE, JOIN, etc.)
# Step 2 – detect label (word(s) + colon) before the LOAD keyword
# Step 3 – locate the LOAD keyword; everything up to FROM/RESIDENT is fields
# Step 4 – everything after FROM/RESIDENT is the source (clean it)

_LOAD_KW_RE = re.compile(
    r'\b(?:LOAD(?:\s+DISTINCT)?|SELECT)\b',
    re.IGNORECASE,
)

_FROM_KW_RE  = re.compile(r'\bFROM\b',     re.IGNORECASE)
_RES_KW_RE   = re.compile(r'\bRESIDENT\b', re.IGNORECASE)

# A valid table label is one or more words/brackets, NOT a keyword itself
_LOAD_KEYWORDS = {'LOAD','SELECT','DISTINCT','NOCONCATENATE','CONCATENATE',
                  'JOIN','KEEP','INNER','LEFT','RIGHT','FROM','RESIDENT',
                  'WHERE','IF','THEN','ELSE','END','AND','OR','NOT',
                  'STORE','DROP','SET','LET','TRACE','EXIT','CALL','DO',
                  'WHILE','LOOP','FOR','NEXT','SUB'}


def looks_like_label(text):
    """Return True if text could be a table label (not a field expression)."""
    text = text.strip()
    if not text:
        return False
    # If it starts with a function call pattern like IF( or LEN( → not a label
    if re.match(r'^\w+\s*\(', text):
        return False
    # If it contains operators that indicate a field expression
    if re.search(r'[+\-*/%<>=!&|]', text):
        return False
    # Pure word characters, spaces, brackets — likely a label
    if re.match(r'^[\w\s\[\]]+$', text):
        first_word = text.split()[0].upper()
        return first_word not in _LOAD_KEYWORDS
    return False


def parse_statement(tab, stmt):
    stmt = re.sub(r'\s+', ' ', stmt).strip()
    rows = []

    # ── Step 1: strip leading prefixes, note if JOIN present ──────────────────
    prefix_raw = ''
    m = _ALL_PREFIXES.match(stmt)
    if m:
        prefix_raw = m.group(0)
        stmt = stmt[m.end():].strip()

    join_label = extract_join_label(prefix_raw.strip())

    # ── Step 2: find the LOAD keyword ─────────────────────────────────────────
    load_m = _LOAD_KW_RE.search(stmt)
    if not load_m:
        return []   # not a LOAD/SELECT statement

    before_load = stmt[:load_m.start()].strip()  # might be "TableName:" or ""
    after_load  = stmt[load_m.end():].strip()    # "field1, field2 FROM source"

    # ── Step 3: determine table name ──────────────────────────────────────────
    # before_load may be "FinalTable: NoConcatenate" or "TableName:" or ""
    # Use the first colon as the label boundary.
    table_name = None
    if ':' in before_load:
        colon_idx = before_load.index(':')
        candidate = before_load[:colon_idx].strip()
        candidate = strip_prefixes(candidate).strip()
        if looks_like_label(candidate):
            table_name = candidate

    if table_name is None:
        # unlabelled — use join label or [Unnamed LOAD]
        table_name = join_label if join_label else '[Unnamed LOAD]'

    # ── Step 4: split fields from FROM/RESIDENT ───────────────────────────────
    # Find the last FROM or RESIDENT that is at depth-0 (outside parens/strings)
    from_pos = _find_keyword_pos(after_load, ('FROM', 'RESIDENT'))
    if from_pos is None:
        return []

    fields_raw, kw_and_source = after_load[:from_pos].strip(), after_load[from_pos:].strip()

    # Determine if it's FROM or RESIDENT
    is_resident = kw_and_source.upper().startswith('RESIDENT')
    source_raw  = kw_and_source.split(None, 1)[1].strip() if ' ' in kw_and_source else ''

    from_src = source_raw if is_resident else clean_from(source_raw)
    # For RESIDENT, just take the table name (first word)
    if is_resident:
        from_src = from_src.split()[0] if from_src else ''

    # ── Step 5: emit one row per field ────────────────────────────────────────
    for f in split_fields(fields_raw):
        rows.append({'tab': tab, 'table': table_name, 'field': f, 'from': from_src})

    return rows


def _find_keyword_pos(text, keywords):
    """
    Find the position of the last occurrence of any keyword (word-boundary)
    that is at parenthesis depth 0 and outside string literals.
    Returns the start index or None.
    """
    depth, in_str, qchar, i = 0, False, None, 0
    last_pos = None
    while i < len(text):
        ch = text[i]
        if in_str:
            if ch == qchar:
                in_str = False
        elif ch in ("'", '"'):
            in_str, qchar = True, ch
        elif ch == '(':
            depth += 1
        elif ch == ')':
            depth -= 1
        elif depth == 0:
            for kw in keywords:
                if text[i:i+len(kw)].upper() == kw:
                    # ensure word boundary
                    after = i + len(kw)
                    if after >= len(text) or not text[after].isalnum():
                        before = i - 1
                        if before < 0 or not text[before].isalnum():
                            last_pos = i
                            break
        i += 1
    return last_pos


# ─── pipeline ─────────────────────────────────────────────────────────────────

def extract(qvs_path):
    script = Path(qvs_path).read_text(encoding='utf-8-sig', errors='replace')
    rows = []
    for tab, stmt in tokenise(script):
        rows.extend(parse_statement(tab, stmt))
    return rows


# ─── Excel writer ─────────────────────────────────────────────────────────────

PALETTE = [
    "DEEAF1","E2EFDA","FFF2CC","FCE4D6",
    "EAD1DC","D9D2E9","D6E4F0","FDE9D9",
    "E8F5E9","F3E5F5","E0F7FA","FFF8E1",
]
TAB_COLORS = {}


def tab_fill(name):
    if name not in TAB_COLORS:
        c = PALETTE[len(TAB_COLORS) % len(PALETTE)]
        TAB_COLORS[name] = PatternFill("solid", start_color=c, end_color=c)
    return TAB_COLORS[name]


def write_excel(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "QVS Extraction"

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Tab Name", "Table Name", "Field", "FROM"]
    hfill   = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    hfont   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont; c.fill = hfill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24

    dfont = Font(name="Calibri", size=9)
    for ri, row in enumerate(rows, 2):
        fill = tab_fill(row['tab'])
        for ci, val in enumerate([row['tab'], row['table'], row['field'], row['from']], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = dfont; c.fill = fill; c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[ri].height = 28

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 62
    ws.column_dimensions["D"].width = 48
    ws.freeze_panes = "A2"

    lg = wb.create_sheet("Legend")
    lg["A1"] = "Tab Colour Legend"
    lg["A1"].font = Font(name="Calibri", bold=True, size=12)
    for i, (tab, fill) in enumerate(TAB_COLORS.items(), 3):
        c = lg.cell(row=i, column=1, value=tab)
        c.fill = fill; c.font = Font(name="Calibri", size=10)
        c.border = border; c.alignment = Alignment(vertical="center")
        lg.row_dimensions[i].height = 18
    lg.column_dimensions["A"].width = 40

    wb.save(out_path)
    print(f"[OK] Saved {len(rows)} rows → {out_path}")


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(1)
    qvs_path = sys.argv[1]
    if not Path(qvs_path).exists():
        print(f"[ERROR] File not found: {qvs_path}"); sys.exit(1)
    out_path = (
        sys.argv[2] if len(sys.argv) >= 3
        else Path(qvs_path).stem + "_extracted.xlsx"
    )
    print(f"[INFO] Parsing: {qvs_path}")
    rows = extract(qvs_path)
    print(f"[INFO] Found {len(rows)} rows across {len({r['tab'] for r in rows})} tab(s)")
    if not rows:
        print("[WARN] No LOAD/SELECT statements found."); sys.exit(0)
    write_excel(rows, out_path)


if __name__ == "__main__":
    main()
