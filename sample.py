“””
qvs_to_l3.py

Parses a QVS file and populates the “Level 3 Transformations” tab
in an existing Excel workbook, or creates a new one if not provided.

Usage:
python qvs_to_l3.py –qvs Change2.qvs –excel Stability_Dashboard_updated.xlsx
python qvs_to_l3.py –qvs Change2.qvs   # creates new output Excel
“””

import re
import argparse
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# —————————————————————————

# QVS PARSER

# —————————————————————————

def clean_field_expr(raw: str) -> str:
“”“Strip trailing comma and whitespace from a field expression.”””
return raw.strip().rstrip(”,”).strip()

def extract_alias(field_expr: str):
“””
Given ‘some_expr AS [Alias]’ or ‘some_expr as alias’,
return (expression, alias). If no alias, return (expr, expr).
“””
m = re.match(r”(.+?)\s+[Aa][Ss]\s+([?.+?]?)\s*$”, field_expr, re.DOTALL)
if m:
return m.group(1).strip(), m.group(2).strip(”[]”).strip()
return field_expr, field_expr

def parse_from_clause(block: str):
“”“Extract the source path from a FROM … (qvd) or INLINE block.”””
from_match = re.search(
r’\bFROM\b\s+(.*?)(?:\s*(qvd)|\s*;|\s*WHERE\b|\s*$)’,
block, re.IGNORECASE | re.DOTALL
)
if from_match:
src = from_match.group(1).strip().strip(’”’).strip(”’”)
return src
if re.search(r’\bINLINE\b’, block, re.IGNORECASE):
return “INLINE”
return “”

def parse_qvs(qvs_text: str):
“””
Parse QVS script and return list of dicts:
{tab, table_name, field, from_source, level}
“””
rows = []
current_tab = “Main”

```
# Split into logical blocks by semicolons, preserving structure
# First pass: identify tab markers
tab_pattern = re.compile(r'///\$tab\s+(.+)', re.IGNORECASE)

# We'll process line by line to track tabs, then regex-extract LOAD blocks
lines = qvs_text.splitlines()
tab_map = {}  # line_number -> tab_name
for i, line in enumerate(lines):
    m = tab_pattern.search(line)
    if m:
        tab_map[i] = m.group(1).strip()

# Build a version of text with tab annotations embedded as comments we can split on
annotated = qvs_text
# Replace ///$ tab markers so we can split on them
annotated = re.sub(r'///\$tab\s+(.+)', r'<<<TAB:\1>>>', annotated, flags=re.IGNORECASE)

segments = re.split(r'<<<TAB:(.+?)>>>', annotated)
# segments alternates: [pre_first_tab_text, tab1_name, tab1_body, tab2_name, tab2_body, ...]

tab_sections = []
if segments[0].strip():
    tab_sections.append(("Main", segments[0]))
for i in range(1, len(segments), 2):
    tab_name = segments[i].strip()
    body = segments[i + 1] if i + 1 < len(segments) else ""
    tab_sections.append((tab_name, body))

for tab_name, body in tab_sections:
    rows.extend(parse_tab_body(tab_name, body))

return rows
```

def parse_tab_body(tab_name: str, body: str):
“”“Extract all LOAD statements from a tab body and return L3 rows.”””
rows = []

```
# Split body into semicolon-terminated statements
# We use a simple state machine to handle brackets/strings
statements = split_statements(body)

for stmt in statements:
    stmt = stmt.strip()
    if not stmt:
        continue
    # Skip pure comments
    if stmt.startswith("//"):
        continue

    load_rows = parse_load_statement(tab_name, stmt)
    rows.extend(load_rows)

return rows
```

def split_statements(text: str):
“”“Split QVS text on semicolons, respecting bracket/string contexts.”””
stmts = []
current = []
depth = 0
in_str = False
str_char = None
i = 0
while i < len(text):
c = text[i]
if in_str:
current.append(c)
if c == str_char:
in_str = False
elif c in (’”’, “’”):
in_str = True
str_char = c
current.append(c)
elif c == ‘[’:
depth += 1
current.append(c)
elif c == ‘]’:
depth = max(0, depth - 1)
current.append(c)
elif c == ‘;’ and depth == 0:
stmts.append(””.join(current))
current = []
else:
current.append(c)
i += 1
if current:
stmts.append(””.join(current))
return stmts

def parse_load_statement(tab_name: str, stmt: str):
“””
Given a single QVS statement (without the trailing semicolon),
extract table name, fields, source, and return L3 rows.
“””
rows = []

```
# Detect table name patterns:
# 1. TableName: \n LOAD ...
# 2. LEFT JOIN (TableName) \n LOAD
# 3. CONCATENATE (TableName) \n LOAD
# 4. Mapping LOAD / LOAD DISTINCT ...

# Strip inline comments (// to end of line) — but keep commented-out fields
# for skipping later
stmt_clean = re.sub(r'(?m)^[ \t]*//.*$', '', stmt)

# Check for LOAD keyword
if not re.search(r'\bLOAD\b', stmt_clean, re.IGNORECASE):
    return rows

# --- Determine table name ---
table_name = ""

# Pattern: TableName:\nLOAD or TableName:\nMapping LOAD
tbl_match = re.match(r'^\s*([A-Za-z_][A-Za-z0-9_ ]*?):\s*\n', stmt_clean)
if tbl_match:
    table_name = tbl_match.group(1).strip()

# Pattern: LEFT JOIN (TableName) or INNER JOIN (TableName)
join_match = re.search(r'\b(?:LEFT|INNER|OUTER|RIGHT)?\s*JOIN\s*\(([^)]+)\)', stmt_clean, re.IGNORECASE)
if join_match:
    table_name = join_match.group(1).strip()

# Pattern: CONCATENATE (TableName)
concat_match = re.search(r'\bCONCATENATE\s*\(([^)]+)\)', stmt_clean, re.IGNORECASE)
if concat_match and not table_name:
    table_name = concat_match.group(1).strip() + " (CONCATENATE)"
elif concat_match:
    table_name = table_name + " (CONCATENATE)"

# Pattern: [Table Name]: Load
bracket_tbl = re.match(r'^\s*\[([^\]]+)\]:\s*\n', stmt_clean)
if bracket_tbl and not table_name:
    table_name = bracket_tbl.group(1).strip()

if not table_name:
    table_name = "(unnamed)"

# --- Determine source ---
from_source = parse_from_clause(stmt_clean)

# --- Extract fields between LOAD and FROM/WHERE/; ---
# Find the LOAD keyword position
load_match = re.search(r'\bLOAD\b\s*(DISTINCT\s*)?', stmt_clean, re.IGNORECASE)
if not load_match:
    return rows

fields_start = load_match.end()

# Find end of field list: FROM keyword or end of string
from_pos_match = re.search(r'\bFROM\b', stmt_clean[fields_start:], re.IGNORECASE)
inline_pos_match = re.search(r'\bINLINE\b', stmt_clean[fields_start:], re.IGNORECASE)

if from_pos_match:
    fields_end = fields_start + from_pos_match.start()
elif inline_pos_match:
    fields_end = fields_start + inline_pos_match.start()
else:
    fields_end = len(stmt_clean)

fields_block = stmt_clean[fields_start:fields_end].strip()

# Split field block by commas (respecting nested parens/brackets)
fields = split_field_list(fields_block)

for field_raw in fields:
    field_raw = clean_field_expr(field_raw)
    if not field_raw:
        continue
    # Skip lines that are pure comments
    if field_raw.startswith("//"):
        continue

    expr, alias = extract_alias(field_raw)

    rows.append({
        "tab": tab_name,
        "table_name": table_name,
        "field": alias,
        "expression": expr,
        "from": from_source,
        "level": "Level 3"
    })

return rows
```

def split_field_list(text: str):
“”“Split a comma-separated field list respecting nested parens.”””
fields = []
current = []
depth = 0
for c in text:
if c == ‘(’:
depth += 1
current.append(c)
elif c == ‘)’:
depth -= 1
current.append(c)
elif c == ‘,’ and depth == 0:
fields.append(””.join(current).strip())
current = []
else:
current.append(c)
if current:
fields.append(””.join(current).strip())
return fields

# —————————————————————————

# EXCEL WRITER

# —————————————————————————

HEADER = [“Tab Name”, “Table Name”, “Field”, “From”, “Level”]
SHEET_NAME = “Level 3 Transformations”

HDR_FILL = PatternFill(“solid”, start_color=“1F4E79”, end_color=“1F4E79”)
HDR_FONT = Font(bold=True, color=“FFFFFF”, name=“Arial”, size=10)
DATA_FONT = Font(name=“Arial”, size=10)
ALT_FILL = PatternFill(“solid”, start_color=“D9E1F2”, end_color=“D9E1F2”)
WHITE_FILL = PatternFill(“solid”, start_color=“FFFFFF”, end_color=“FFFFFF”)
THIN = Side(style=“thin”, color=“BFBFBF”)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
COL_WIDTHS = [20, 35, 45, 65, 10]

def write_excel(rows, excel_path=None, output_path=“L3_Transformations_output.xlsx”):
if excel_path and Path(excel_path).exists():
wb = load_workbook(excel_path)
if SHEET_NAME in wb.sheetnames:
ws = wb[SHEET_NAME]
# Find last used row
last_row = ws.max_row
# Check if sheet has header
if ws.cell(1, 1).value != “Tab Name”:
# Write header
_write_header(ws, 1)
start_row = 2
else:
start_row = last_row + 1
else:
ws = wb.create_sheet(SHEET_NAME)
_write_header(ws, 1)
_set_col_widths(ws)
start_row = 2
else:
wb = Workbook()
ws = wb.active
ws.title = SHEET_NAME
_write_header(ws, 1)
_set_col_widths(ws)
start_row = 2

```
for i, row in enumerate(rows):
    r = start_row + i
    fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
    values = [
        row["tab"],
        row["table_name"],
        row["field"],
        row["from"],
        row["level"]
    ]
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = DATA_FONT
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=False, vertical="center")

# Freeze header row
ws.freeze_panes = "A2"

wb.save(output_path)
print(f"✅ Written {len(rows)} rows to '{output_path}' → sheet '{SHEET_NAME}'")
```

def _write_header(ws, row):
for c, col_name in enumerate(HEADER, 1):
cell = ws.cell(row=row, column=c, value=col_name)
cell.font = HDR_FONT
cell.fill = HDR_FILL
cell.border = BORDER
cell.alignment = Alignment(horizontal=“center”, vertical=“center”)

def _set_col_widths(ws):
for i, width in enumerate(COL_WIDTHS, 1):
ws.column_dimensions[get_column_letter(i)].width = width

# —————————————————————————

# MAIN

# —————————————————————————

def main():
parser = argparse.ArgumentParser(description=“Parse QVS → L3 Transformations Excel”)
parser.add_argument(”–qvs”, required=True, help=“Path to the .qvs file”)
parser.add_argument(”–excel”, default=None, help=“Existing Excel file to update (optional)”)
parser.add_argument(”–output”, default=“L3_Transformations_output.xlsx”, help=“Output file path”)
parser.add_argument(”–tab”, default=None, help=“Filter to a specific QVS tab name (e.g. ‘Incident’)”)
args = parser.parse_args()

```
qvs_text = Path(args.qvs).read_text(encoding="utf-8-sig", errors="replace")
print(f"📄 Parsing {args.qvs} ({len(qvs_text)} chars) ...")

rows = parse_qvs(qvs_text)
print(f"🔍 Found {len(rows)} field rows across all tabs")

if args.tab:
    rows = [r for r in rows if r["tab"].lower() == args.tab.lower()]
    print(f"🔎 Filtered to tab '{args.tab}': {len(rows)} rows")

if not rows:
    print("⚠️  No rows to write. Check your QVS file or tab name filter.")
    return

write_excel(rows, excel_path=args.excel, output_path=args.output)
```

if **name** == “**main**”:
main()