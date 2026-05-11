"""
QVD Name → QVW Search → Table Extraction → Excel Population
============================================================
Reads QVD names from Column D of the "Lineage Structure" sheet,
searches a folder for QVW files that reference each QVD,
parses all table names from the QVW's load script,
and writes results back into the Excel file.

USAGE:
    pip install openpyxl
    python qvd_mapper.py

CONFIGURE the three paths at the top of main() before running.
"""

import os
import re
import zipfile
import struct
import openpyxl
from pathlib import Path


# ─────────────────────────────────────────────
# 1. QVW TEXT EXTRACTION
# QVW files are binary but embed the load script as UTF-16-LE text.
# We search for the script section inside the raw bytes.
# ─────────────────────────────────────────────

def extract_script_from_qvw(qvw_path: Path) -> str:
    """
    Extract the load script text from a QVW binary file.
    QVW stores the script as UTF-16-LE. We scan for the
    pattern that precedes it and decode until a null section.
    Falls back to a broad regex scan if the header isn't found.
    """
    try:
        raw = qvw_path.read_bytes()

        # QVW stores script blocks marked by the tag b'ESCRIPT\x00' or similar.
        # Broad approach: decode whole file as UTF-16-LE, ignoring errors.
        try:
            text_utf16 = raw.decode("utf-16-le", errors="ignore")
        except Exception:
            text_utf16 = ""

        # Also try UTF-8 decode (some newer QVW/QVD have UTF-8 script sections)
        try:
            text_utf8 = raw.decode("utf-8", errors="ignore")
        except Exception:
            text_utf8 = ""

        # Combine both — we'll run our regex on whichever has more SQL keywords
        combined = text_utf16 if text_utf16.count("LOAD") > text_utf8.count("LOAD") else text_utf8
        return combined

    except Exception as e:
        print(f"  [WARN] Could not read {qvw_path.name}: {e}")
        return ""


# ─────────────────────────────────────────────
# 2. TABLE NAME PARSING FROM QLIK LOAD SCRIPT
# ─────────────────────────────────────────────

def parse_table_names(script_text: str) -> list[str]:
    """
    Extract table names from a Qlik load script.

    Qlik table names appear in patterns like:
        TableName:
        LOAD ...
        FROM ...;

        [Table Name]:
        LOAD ...;

    Also catches SQL SELECT FROM <table> patterns.
    """
    table_names = set()

    # Pattern 1: Labelled LOAD  →  TableName:\nLOAD
    labelled_load = re.findall(
        r'^\s*\[?([A-Za-z0-9_ ]+?)\]?\s*:\s*\n\s*(?:LOAD|SELECT)',
        script_text, re.MULTILINE | re.IGNORECASE
    )
    table_names.update(t.strip() for t in labelled_load)

    # Pattern 2: FROM RTIM."TABLE_NAME"  (SQL source tables)
    sql_from = re.findall(
        r'\bFROM\b\s+(?:\w+\.)?["\[]?([A-Za-z0-9_]+)["\]]?',
        script_text, re.IGNORECASE
    )
    table_names.update(t.strip() for t in sql_from if len(t) > 2)

    # Pattern 3: Inline variable expansions like $(vTable1) already resolved
    # to table names — skip variable references (they start with $)
    table_names = {t for t in table_names if not t.startswith("$") and not t.startswith("v")}

    # Remove common noise words that aren't table names
    noise = {"LOAD", "SELECT", "WHERE", "WITH", "UR", "FROM", "AND", "OR", "NOT", "AS", "BY"}
    table_names -= noise

    return sorted(table_names)


# ─────────────────────────────────────────────
# 3. QVD SEARCH INSIDE SCRIPT
# ─────────────────────────────────────────────

def script_creates_qvd(script_text: str, qvd_name: str) -> bool:
    """
    Check if a QVW load script GENERATES/CREATES the given QVD file.

    In Qlik, a QVW produces a QVD via the STORE statement:
        STORE TableName INTO $(vQVDExtractPath)STABILITY_Change.qvd;
        CALL StoreQVD('STABILITY_Change', ...)   <- subroutine wrapper pattern

    We match three patterns:
      1. STORE ... INTO ... <qvd_name>.qvd
      2. CALL StoreQVD ... <qvd_name>   (common wrapper seen in Image 3)
      3. Fallback: filename appears anywhere + STORE exists in script
    """
    base = qvd_name.replace(".qvd", "").strip()
    escaped = re.escape(base)

    # Pattern 1: STORE <table> INTO <path><qvd_name>.qvd
    store_pattern = rf'\bSTORE\b[^;]*\bINTO\b[^;]*{escaped}[^;]*\.qvd'
    if re.search(store_pattern, script_text, re.IGNORECASE):
        return True

    # Pattern 2: CALL StoreQVD ('TableName', 'qvd_name', ...)
    call_pattern = rf'\bCALL\b\s+StoreQVD[^;]*{escaped}'
    if re.search(call_pattern, script_text, re.IGNORECASE):
        return True

    # Pattern 3: Fallback — filename appears anywhere + STORE exists in script
    # Catches variable-wrapped paths like $(vExtractFolder)STABILITY_Change.qvd
    if re.search(escaped, script_text, re.IGNORECASE):
        if re.search(r'\bSTORE\b', script_text, re.IGNORECASE):
            return True

    return False


# ─────────────────────────────────────────────
# 4. FOLDER SCAN FOR QVW FILES
# ─────────────────────────────────────────────

def find_qvw_files(folder: str | Path) -> list[Path]:
    """Recursively find all .qvw files under folder."""
    folder = Path(folder)
    qvw_files = list(folder.rglob("*.qvw"))
    print(f"Found {len(qvw_files)} QVW file(s) in {folder}")
    return qvw_files


# ─────────────────────────────────────────────
# 5. EXCEL READ + WRITE
# ─────────────────────────────────────────────

def read_qvd_names_from_excel(excel_path: str | Path, sheet_name: str = "Lineage Structure") -> dict[int, str]:
    """
    Read the 'Final Dashboard QVD Name' column from the 'Lineage Structure' sheet.

    Sheet layout:
      Row 1 → group headers  : Database | Extract Level 1 | Transformation Level 2 | Data model Level 3 | Data Model
      Row 2 → column headers : Database Table Name | Extract QVD Name | Transform QVD Name | Final Dashboard QVD Name | Table Name
      Row 3+ → data

    Dynamically locates the column by searching for the header text
    "Final Dashboard QVD Name" in row 2, so it works even if columns shift.
    Returns {excel_row_number: qvd_name}
    """
    wb = openpyxl.load_workbook(excel_path)

    # Validate sheet exists
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found.\n"
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[sheet_name]

    # ── Dynamically find the column with header "Final Dashboard QVD Name" ──
    TARGET_HEADER = "Final Dashboard QVD Name"
    qvd_col = None

    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    for col_idx, cell_val in enumerate(header_row, start=1):
        if cell_val and str(cell_val).strip().lower() == TARGET_HEADER.lower():
            qvd_col = col_idx
            break

    if qvd_col is None:
        # Print what row 2 actually contains to help the user debug
        found_headers = [str(v) for v in header_row if v]
        raise ValueError(
            f"Could not find column '{TARGET_HEADER}' in row 2 of sheet '{sheet_name}'.\n"
            f"Row 2 contains: {found_headers}"
        )

    col_letter = openpyxl.utils.get_column_letter(qvd_col)
    print(f"Found '{TARGET_HEADER}' in column {col_letter} (col index {qvd_col})")

    # ── Read data from row 3 downward ──
    qvd_map = {}
    for row_num in range(3, ws.max_row + 1):
        val = ws.cell(row=row_num, column=qvd_col).value
        if val and str(val).strip():
            qvd_map[row_num] = str(val).strip()

    print(f"Read {len(qvd_map)} QVD name(s) from '{sheet_name}' column {col_letter} (rows 3–{ws.max_row})")
    return qvd_map, wb, ws


def write_results_to_excel(
    wb: openpyxl.Workbook,
    ws,
    results: dict[int, dict],
    output_path: str | Path,
    qvw_col: int = 6,      # Column F — QVW file name(s)
    tables_col: int = 7,   # Column G — extracted table names
):
    """
    Write QVW filename and table names back into the Excel sheet.
    qvw_col and tables_col are 1-based column indices.
    """
    # Write headers if not present
    if not ws.cell(row=2, column=qvw_col).value:
        ws.cell(row=2, column=qvw_col).value = "QVW File(s)"
    if not ws.cell(row=2, column=tables_col).value:
        ws.cell(row=2, column=tables_col).value = "Source Table Names"

    for row_num, info in results.items():
        ws.cell(row=row_num, column=qvw_col).value = info.get("qvw_files", "")
        ws.cell(row=row_num, column=tables_col).value = info.get("table_names", "")

    wb.save(output_path)
    print(f"\nResults written to: {output_path}")


# ─────────────────────────────────────────────
# 6. MAIN ORCHESTRATION
# ─────────────────────────────────────────────

def main():
    # ── CONFIGURE THESE THREE PATHS ──────────────────────────────────
    EXCEL_PATH   = r"C:\path\to\Stability_Dashboard.xlsx"   # your Excel file
    SEARCH_FOLDER = r"\\vmwas1388330\QVDOCS\QAPM\QVDATA\Model"  # QVW folder
    OUTPUT_PATH  = r"C:\path\to\Stability_Dashboard_updated.xlsx"  # output
    SHEET_NAME   = "Lineage Structure"
    # ─────────────────────────────────────────────────────────────────

    print("=" * 60)
    print("QVD → QVW → Table Names → Excel Mapper")
    print("=" * 60)

    # Step 1: Read QVD names from Excel Column D
    qvd_map, wb, ws = read_qvd_names_from_excel(EXCEL_PATH, SHEET_NAME)

    # Step 2: Scan folder for all QVW files and pre-read their scripts
    qvw_files = find_qvw_files(SEARCH_FOLDER)

    print("\nPre-reading QVW scripts (this may take a moment)...")
    qvw_scripts = {}  # {Path: script_text}
    for qvw in qvw_files:
        script = extract_script_from_qvw(qvw)
        qvw_scripts[qvw] = script
        print(f"  ✓ {qvw.name} ({len(script):,} chars decoded)")

    # Step 3: For each QVD name, find matching QVW files and extract tables
    results = {}  # {row_num: {"qvw_files": str, "table_names": str}}

    print("\nMatching QVD names to QVW scripts...")
    for row_num, qvd_name in qvd_map.items():
        matching_qvws = []
        all_tables = set()

        for qvw_path, script_text in qvw_scripts.items():
            if script_creates_qvd(script_text, qvd_name):
                matching_qvws.append(qvw_path.name)
                tables = parse_table_names(script_text)
                all_tables.update(tables)

        qvw_str    = " | ".join(matching_qvws) if matching_qvws else "NOT FOUND"
        tables_str = ", ".join(sorted(all_tables)) if all_tables else ""

        results[row_num] = {
            "qvw_files":   qvw_str,
            "table_names": tables_str,
        }

        status = f"  Row {row_num}: {qvd_name}"
        if matching_qvws:
            print(f"{status} → {qvw_str} | Tables: {tables_str[:80]}{'...' if len(tables_str)>80 else ''}")
        else:
            print(f"{status} → [no match]")

    # Step 4: Write results back to Excel
    write_results_to_excel(wb, ws, results, OUTPUT_PATH)

    print("\nDone.")


if __name__ == "__main__":
    main()
