## “””
QVW Log Path Extractor

Reads a QVW .log file and extracts the resolved file path for each
Qlik variable (e.g. APTDDCFile, AlertsFile, etc.), then writes the
results into an Excel file with two columns: Variable Name | QVD Path.

Usage:
1. Set LOG_FILE to your .log file path
2. Set OUTPUT_EXCEL to where you want the output saved
3. Optionally set EXISTING_EXCEL if you want to write into your
existing spreadsheet instead of creating a new one
4. Run: python extract_qvw_paths.py
“””

import re
import os

# ── Configuration ─────────────────────────────────────────────────────────────

LOG_FILE       = r”TORG_ETL.QVW.log”        # Path to your .log file
OUTPUT_EXCEL   = r”QVD_Paths_Output.xlsx”   # Output Excel file path

# If you want to write into your existing Excel (TORG.xlsx), set this.

# The script will fill column B (QVD Path) next to existing variable names.

# Set to None to just create a fresh output file.

EXISTING_EXCEL = None   # e.g. r”TORG.xlsx”
EXISTING_SHEET = “QVD List”
EXISTING_COL_VARS = “A”   # Column with variable names like $(APTDDCFile)
EXISTING_COL_PATH = “B”   # Column to write paths into
EXISTING_START_ROW = 2    # First data row (skip header)

# ── Parse Log File ─────────────────────────────────────────────────────────────

def parse_log(log_path):
“””
Extracts {VariableName: FilePath} from the log.

```
The log pattern is:
    Let [VarName] = Peek('FileLocation', vI, 'FileLocations')
    ...
    SET vFilePath = <actual path>
"""
with open(log_path, "r", encoding="utf-8", errors="replace") as f:
    content = f.read()

# Match blocks: Let [VarName] ... SET vFilePath = <path>
# The path continues until end of line
pattern = re.compile(
    r"Let\s+\[([^\]]+)\]\s*=\s*Peek\(['\"]FileLocation['\"].*?\n"   # Let [VarName] line
    r"(?:.*?\n)*?"                                                    # any lines in between
    r"SET\s+vFilePath\s*=\s*(.+)",                                   # SET vFilePath = <path>
    re.IGNORECASE
)

results = {}
for match in pattern.finditer(content):
    var_name = match.group(1).strip()
    file_path = match.group(2).strip()
    results[var_name] = file_path

return results
```

def parse_log_fallback(log_path):
“””
Line-by-line fallback parser — more robust if the regex above
misses entries due to log formatting variations.

```
Tracks the most recently seen Let [VarName] assignment and pairs
it with the next SET vFilePath line.
"""
results = {}
current_var = None

let_pattern      = re.compile(r"Let\s+\[([^\]]+)\]\s*=\s*Peek\(['\"]FileLocation['\"]", re.IGNORECASE)
filepath_pattern = re.compile(r"SET\s+vFilePath\s*=\s*(.+)", re.IGNORECASE)

with open(log_path, "r", encoding="utf-8", errors="replace") as f:
    for line in f:
        let_match = let_pattern.search(line)
        if let_match:
            current_var = let_match.group(1).strip()

        if current_var:
            fp_match = filepath_pattern.search(line)
            if fp_match:
                results[current_var] = fp_match.group(1).strip()
                current_var = None   # reset after pairing

return results
```

# ── Write Output ───────────────────────────────────────────────────────────────

def write_fresh_excel(path_map, output_path):
“”“Creates a new Excel file with Variable | Path columns.”””
try:
import openpyxl
except ImportError:
import subprocess, sys
subprocess.check_call([sys.executable, “-m”, “pip”, “install”, “openpyxl”, “-q”])
import openpyxl

```
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "QVD Paths"

# Header
ws["A1"] = "Variable Name"
ws["B1"] = "Resolved Path"
ws["A1"].font = openpyxl.styles.Font(bold=True)
ws["B1"].font = openpyxl.styles.Font(bold=True)

for row_idx, (var, path) in enumerate(sorted(path_map.items()), start=2):
    ws.cell(row=row_idx, column=1, value=var)
    ws.cell(row=row_idx, column=2, value=path)

# Auto-width
for col in ws.columns:
    max_len = max((len(str(c.value)) for c in col if c.value), default=10)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

wb.save(output_path)
print(f"  Saved fresh output → {output_path}")
```

def write_into_existing_excel(path_map, excel_path, sheet_name, col_vars, col_path, start_row):
“””
Reads variable names from an existing Excel file and fills in
the resolved path in the adjacent column.
“””
try:
import openpyxl
except ImportError:
import subprocess, sys
subprocess.check_call([sys.executable, “-m”, “pip”, “install”, “openpyxl”, “-q”])
import openpyxl

```
wb = openpyxl.load_workbook(excel_path)
ws = wb[sheet_name]

col_vars_idx = openpyxl.utils.column_index_from_string(col_vars)
col_path_idx = openpyxl.utils.column_index_from_string(col_path)

matched = 0
unmatched = []

for row in ws.iter_rows(min_row=start_row):
    cell_val = row[col_vars_idx - 1].value
    if not cell_val:
        continue

    # Strip $() wrapper if present: $(APTDDCFile) → APTDDCFile
    var_name = re.sub(r"^\$\(|\)$", "", str(cell_val).strip())

    if var_name in path_map:
        ws.cell(row=row[0].row, column=col_path_idx, value=path_map[var_name])
        matched += 1
    else:
        unmatched.append(var_name)

wb.save(excel_path)
print(f"  Updated {matched} rows in '{sheet_name}' → {excel_path}")
if unmatched:
    print(f"  No path found for {len(unmatched)} variables:")
    for u in unmatched:
        print(f"    - {u}")
```

# ── Main ───────────────────────────────────────────────────────────────────────

def main():
if not os.path.exists(LOG_FILE):
print(f”ERROR: Log file not found → {LOG_FILE}”)
return

```
print(f"Parsing log file: {LOG_FILE}")

# Try regex parser first, fall back to line-by-line if it finds nothing
path_map = parse_log(LOG_FILE)
if not path_map:
    print("  Regex parser found nothing, trying line-by-line fallback...")
    path_map = parse_log_fallback(LOG_FILE)

if not path_map:
    print("  No variable→path mappings found. Check log format.")
    return

print(f"  Found {len(path_map)} variable→path mappings.")

if EXISTING_EXCEL and os.path.exists(EXISTING_EXCEL):
    print(f"Writing into existing Excel: {EXISTING_EXCEL}")
    write_into_existing_excel(
        path_map,
        EXISTING_EXCEL,
        EXISTING_SHEET,
        EXISTING_COL_VARS,
        EXISTING_COL_PATH,
        EXISTING_START_ROW
    )
else:
    print(f"Creating output Excel: {OUTPUT_EXCEL}")
    write_fresh_excel(path_map, OUTPUT_EXCEL)

# Also print to console
print("\n── Extracted Paths ──────────────────────────────────────")
for var, path in sorted(path_map.items()):
    print(f"  {var:<35} → {path}")
```

if **name** == “**main**”:
main()